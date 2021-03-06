import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import string


# Get value from Internet
html = requests.get("https://finance.naver.com/sise/sise_market_sum.nhn?sosok=0") #cospi
soup = BeautifulSoup(html.content, "html.parser")
num = soup.select('td.number')
corp = soup.find_all('a', attrs={'class': 'tltle'})
c = []
for h in corp:
    c.append(str(h.get_text()))
a = []
b = {}
how_first = 0
how = 10
name_number = ['현재가', '전일비', '등락률', '액면가',
               '시가총액', '상장주식', '외국인비율', '거래량', 'PER', 'ROE']
for x in num:
    a.append(str((x.get_text()).strip()))

for j in range(50):
    for i in a[how-10:how]:
        b[c[j]+' '+name_number[how_first]] = i
        how_first += 1
    how_first = 0
    how += 10

# SpreadSheet
wb = Workbook()
ws = wb.active

ws.title = "Main Sheet"

cell = ws['A1']
cell.value = "종목명"

for m in c:  # 종목명 삽입
    cell = ws['A'+str(c.index(m)+2)]
    cell.value = m

alphabet = list(string.ascii_uppercase)

for n in name_number:  # name_number 삽입
    s = name_number.index(n)
    cell = ws[alphabet[s+1]+'1']
    cell.value = n

for i in name_number:  # insert every value on cell
    for j in c:
        s = name_number.index(i)
        cell = ws[alphabet[s+1]+str(c.index(j)+2)]
        cell.value = b[j+' '+i]
    ws.column_dimensions[alphabet[name_number.index(i)]].auto_size = True  # 열 너비 자동 맞춤
wb.save("Test.xlsx")