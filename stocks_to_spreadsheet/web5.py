import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import string
import time


class GetSource_MakeSheet():
    real_pages = "https://finance.naver.com/sise/sise_market_sum.nhn?sosok="
    name_number = ['현재가', '전일비', '등락률', '액면가',
                   '시가총액', '상장주식', '외국인비율', '거래량', 'PER', 'ROE']
    # SpreadSheet
    wb = Workbook()
    ws = wb.active

    ws.title = "Main Sheet"

    cell = ws['A1']
    cell.value = "종목명"
    a = []
    b = {}
    c = []

    def __init__(self, cos, pages):
        if cos == 1:
            self.real_pages += "0&page="  # cospi
        elif cos == 2:
            self.real_pages += "1&page="  # cosdaq
        else:
            print("invalid value")
            return

        for i in range(pages):
            self.parsing_pages(i+1)
        today = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        self.wb.save(today+".xlsx")

    def parsing_pages(self, pages):
        nowpage = requests.get(self.real_pages + str(pages))
        # Get value from Internet
        self.soup = BeautifulSoup(nowpage.content, "html.parser")
        self.num = self.soup.select('td.number')
        self.corp = self.soup.find_all('a', attrs={'class': 'tltle'})
        self.get50company(self.corp)
        self.makevaluelist(self.num)
        self.makedictionary()
        self.make_spreadsheet(pages)
        self.c = []
        self.a = []

    def get50company(self, companyname):
        for h in companyname:
            self.c.append(str(h.get_text()))  # make company name list

    def makevaluelist(self, num):
        for x in num:  # make number value list
            self.a.append(str((x.get_text()).strip()))

    def makedictionary(self):
        how_first = 0
        how = 10
        for j in range(50):  # make dictionary
            for i in self.a[how-10:how]:
                self.b[self.c[j]+' '+self.name_number[how_first]] = i
                how_first += 1
            how_first = 0
            how += 10

    def make_spreadsheet(self, pages):
        for m in self.c:  # 종목명 삽입
            self.cell = self.ws['A' +
                                str(50*(pages-1)+(self.c.index(m)+2))]  # 2345
            self.cell.value = m

        alphabet = list(string.ascii_uppercase)

        for n in self.name_number:  # name_number 삽입
            s = self.name_number.index(n)
            self.cell = self.ws[alphabet[s+1]+'1']
            self.cell.value = n

        for i in self.name_number:  # insert every value on cell
            for j in self.c:
                s = self.name_number.index(i)
                self.cell = self.ws[alphabet[s+1] +
                                    str(50*(pages-1)+self.c.index(j)+2)]
                self.cell.value = self.b[j+' '+i]
            self.ws.column_dimensions[alphabet[self.name_number.index(
                i)]].auto_size = True  # 열 너비 자동 맞춤

# def GetSource_MakeSheet(cos, pages):

#         # SpreadSheet
#         wb = Workbook()
#         ws = wb.active

#         ws.title = "Main Sheet"

#         cell = ws['A1']
#         cell.value = "종목명"

#         for m in c:  # 종목명 삽입
#             cell = ws['A'+str(50*pages+(c.index(m)+2))]  # 2345
#             cell.value = m

#         alphabet = list(string.ascii_uppercase)

#         for n in name_number:  # name_number 삽입
#             s = name_number.index(n)
#             cell = ws[alphabet[s+1]+'1']
#             cell.value = n

#         for i in name_number:  # insert every value on cell
#             for j in c:
#                 s = name_number.index(i)
#                 cell = ws[alphabet[s+1]+str(c.index(j)+2)]
#                 cell.value = b[j+' '+i]
#             ws.column_dimensions[alphabet[name_number.index(i)]].auto_size = True  # 열 너비 자동 맞춤


#     today = time.strftime('%Y-%m-%d', time.localtime(time.time()))
#     wb.save(today+".xlsx")
data = int(input("Data Source를 선택해주십시오 (1 : 코스피, 2 : 코스닥): "))
page = int(input("몇 페이지까지 확인하시겠습니까?(1페이지 당 50종목): "))
GetSource_MakeSheet(data, page)
