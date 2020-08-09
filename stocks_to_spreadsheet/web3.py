# 날짜 형식 모듈
import datetime
# 엑셀을 사용하기 위한 모듈
from openpyxl import Workbook

# 엑셀 Workbook 생성
wb = Workbook()
# Workbook을 생성하면 기본적으로 시트가 하나 생성됩니다.
print(wb.sheetnames)

# 현재 활성화된 엑셀 시트 객체 가져오기
ws = wb.active
# 시트 이름 바꾸기
ws.title = "Test1"
# 시트 이름으로 시트를 선택할 수도 있습니다.(시트명이 없는 것을 선택하면 에러가 발생합니다.)
ws = wb["Test1"]

# 시트를 생성하기
ws1 = wb.create_sheet(title="Test2")
# 총 시트를 출력하기
print(wb.sheetnames)

# 시트 내에 cell를 선택하는 방법입니다.
cell = ws.cell(row=1, column=1)
# 선택된 cell의 값 넣기
cell.value = "hello world"
# cell의 문자 포멧(general은 excel에서 일반을 뜻합니다.)
print(cell.number_format)

# 2행 1열의 셀을 선택하기.
cell = ws.cell(row=2, column=1)
# 선택된 cell에 날짜 타입의 값 넣기
cell.value = datetime.datetime(2020, 1, 1)
# 셀의 포멧을 변경하기
cell.number_format = 'yyyy-mm-dd'
# cell의 문자 포멧
print(cell.number_format)

# 여기는 Test2의 시트의 값을 넣습니다.
# 셀을 선택하는 방법은 R1C1 형식으로 선택 가능합니다.
cell = ws1["A1"]
# 선택된 cell에 값 넣기
cell.value = 5

cell = ws1["A2"]
cell.value = 6

# 함수식(formulae)은 엑셀 사용하는 방법 그대로 넣으면 됩니다.
ws1["A3"] = "=A1+A2"

# 위 작성한 엑셀 시트를 파일로 저장합니다.
wb.save('Example1.xlsx')
